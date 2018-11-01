#coding=utf-8
import openpyxl
import xlrd
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl import formula
from openpyxl.writer.excel import ExcelWriter
import json

class DataConfig:

    def __init__(self,filename):
        self.filename=filename
        self.wb=load_workbook(self.filename)

    def _get_sheet(self,sheet_name):
        return self.wb.get_sheet_by_name(sheet_name)

    def get_sheets(self):
        return self.wb.get_sheet_names()

    def get_max_row(self,sheet_name):
        """
        :param sheet_name:
        :return:
        """
        sheet=self._get_sheet(sheet_name)
        return sheet.max_row

    def get_max_col(self,sheet_name):
        """
        :param sheet_name:
        :return:
        """
        sheet=self._get_sheet(sheet_name)
        return  sheet.max_column

    def get_TestCase_Parmerner(self,sheet_name,test_case_name=None):
        """
        :param sheet_name: sheet名，也是测试模块名
        :param test_case_name: 测试方法名
        :return: 测试参数列表
        """
        data_dic = {}
        max_row=self.get_max_row(sheet_name)
        max_column=self.get_max_col(sheet_name)
        ws=self._get_sheet(sheet_name)
        for ro in range (1, max_row + 1):
            temp_list = []
            pid = ws.cell (row=ro, column=1).value
            for co in range (2, max_column + 1):
                w = ws.cell (row=ro, column=co).value
                temp_list.append (w)
            data_dic[pid] = temp_list
        return data_dic

if __name__=='__main__':
    d=DataConfig(filename='DataSource.xlsx')
    for i  in d.get_TestCase_Parmerner(sheet_name='LoginInerTest'):
        print (str(i))

